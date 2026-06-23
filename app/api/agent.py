"""测试用例生成 Agent API。"""
import io
import json
import logging
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from langchain_core.messages import AIMessage, ToolMessage
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.agent.testcase_agent import build_testcase_agent
from app.db import models
from app.db.session import get_db
from app.judge.judge import judge_case
from app.judge.judge_graph import generate_with_judge_graph, render_mermaid
from app.schemas.common import Msg, Page
from app.schemas.test_case import (
    AgentGenerateRequest,
    BulkDeleteRequest,
    BulkResult,
    BulkUpdateRequest,
    JudgeGraphDebugRequest,
    ProjectStats,
    TestCaseOut,
    TestCaseUpdate,
)

logger = logging.getLogger(__name__)
router = APIRouter(tags=["agent"])


def _project_or_404(db: Session, project_id: int) -> models.Project:
    obj = db.get(models.Project, project_id)
    if not obj:
        raise HTTPException(404, "Project not found")
    return obj


def _msg_to_event(msg) -> dict | None:
    """把 LangGraph 消息转成 SSE 事件字典，无意义时返回 None。"""
    if isinstance(msg, AIMessage):
        tool_calls = getattr(msg, "tool_calls", None) or []
        if tool_calls:
            return {
                "type": "tool_call",
                "calls": [
                    {"name": tc.get("name"), "args": tc.get("args", {})}
                    for tc in tool_calls
                ],
            }
        content = msg.content if isinstance(msg.content, str) else str(msg.content)
        if content.strip():
            return {"type": "thinking", "content": content}
    elif isinstance(msg, ToolMessage):
        content = msg.content if isinstance(msg.content, str) else str(msg.content)
        return {
            "type": "tool_result",
            "name": getattr(msg, "name", None),
            "content": content[:2000],
        }
    return None


# ---------- 生成接口 ----------

@router.post("/projects/{project_id}/agent/generate")
def agent_generate(
    project_id: int,
    body: AgentGenerateRequest,
    db: Session = Depends(get_db),
):
    """阻塞式执行 Agent，返回完整轨迹 + 最终文本。"""
    _project_or_404(db, project_id)
    agent = build_testcase_agent(project_id)
    result = agent.invoke(
        {"messages": [{"role": "user", "content": body.input}]}
    )
    events = []
    for m in result["messages"]:
        ev = _msg_to_event(m)
        if ev:
            events.append(ev)
    final = ""
    for m in reversed(result["messages"]):
        if isinstance(m, AIMessage) and isinstance(m.content, str) and m.content.strip():
            if not getattr(m, "tool_calls", None):
                final = m.content
                break
    return {"events": events, "final": final}


@router.post("/projects/{project_id}/agent/stream")
def agent_stream(
    project_id: int,
    body: AgentGenerateRequest,
    db: Session = Depends(get_db),
):
    """SSE 流式输出 ReAct 轨迹。"""
    _project_or_404(db, project_id)
    agent = build_testcase_agent(project_id)

    def event_gen():
        seen_ids: set[str] = set()
        try:
            for chunk in agent.stream(
                {"messages": [{"role": "user", "content": body.input}]},
                stream_mode="values",
            ):
                msgs = chunk.get("messages", [])
                if not msgs:
                    continue
                latest = msgs[-1]
                mid = getattr(latest, "id", None) or str(id(latest))
                if mid in seen_ids:
                    continue
                seen_ids.add(mid)
                ev = _msg_to_event(latest)
                if ev is None:
                    continue
                payload = json.dumps(ev, ensure_ascii=False)
                yield f"data: {payload}\n\n"
            yield f"data: {json.dumps({'type': 'done'})}\n\n"
        except Exception as e:
            logger.exception("agent stream error")
            err = json.dumps({"type": "error", "message": str(e)}, ensure_ascii=False)
            yield f"data: {err}\n\n"

    return StreamingResponse(event_gen(), media_type="text/event-stream")


# ---------- 用例 CRUD ----------

@router.get(
    "/projects/{project_id}/test-cases",
    response_model=Page[TestCaseOut],
)
def list_test_cases(
    project_id: int,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=200),
    case_type: str | None = Query(None, description="functional/boundary/exception/negative"),
    priority: str | None = Query(None, description="P0/P1/P2/P3"),
    min_score: int | None = Query(None, ge=0, le=25),
    keyword: str | None = Query(None, description="按标题模糊匹配"),
    db: Session = Depends(get_db),
):
    _project_or_404(db, project_id)
    q = db.query(models.TestCase).filter(
        models.TestCase.project_id == project_id
    )
    if case_type:
        q = q.filter(models.TestCase.case_type == case_type)
    if priority:
        q = q.filter(models.TestCase.priority == priority)
    if min_score is not None:
        q = q.filter(models.TestCase.score >= min_score)
    if keyword:
        q = q.filter(models.TestCase.title.ilike(f"%{keyword}%"))
    total = q.count()
    items = (
        q.order_by(models.TestCase.id.desc())
        .offset((page - 1) * size)
        .limit(size)
        .all()
    )
    return Page[TestCaseOut](items=items, total=total, page=page, size=size)


@router.get("/projects/{project_id}/test-cases/export")
def export_test_cases(
    project_id: int,
    case_type: str | None = Query(None),
    priority: str | None = Query(None),
    min_score: int | None = Query(None, ge=0, le=25),
    keyword: str | None = Query(None),
    db: Session = Depends(get_db),
):
    """按当前过滤条件导出测试用例为 Excel。"""
    project = _project_or_404(db, project_id)
    q = db.query(models.TestCase).filter(models.TestCase.project_id == project_id)
    if case_type:
        q = q.filter(models.TestCase.case_type == case_type)
    if priority:
        q = q.filter(models.TestCase.priority == priority)
    if min_score is not None:
        q = q.filter(models.TestCase.score >= min_score)
    if keyword:
        q = q.filter(models.TestCase.title.ilike(f"%{keyword}%"))
    rows = q.order_by(models.TestCase.id.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "TestCases"
    headers = ["ID", "标题", "类型", "优先级", "前置条件", "操作步骤", "预期结果", "来源", "评分", "创建时间"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="409EFF")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
    widths = [6, 40, 12, 10, 28, 48, 32, 20, 8, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for r in rows:
        steps_text = "\n".join(str(s) for s in (r.steps or []))
        ws.append([
            r.id,
            r.title or "",
            r.case_type or "",
            r.priority or "",
            r.preconditions or "",
            steps_text,
            r.expected or "",
            r.source or "",
            r.score if r.score is not None else "",
            r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "",
        ])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = wrap

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{project.name}_testcases.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"testcases.xlsx\"; filename*=UTF-8''{encoded}"
        },
    )


@router.post("/test-cases/bulk-delete", response_model=BulkResult)
def bulk_delete_test_cases(body: BulkDeleteRequest, db: Session = Depends(get_db)):
    n = (
        db.query(models.TestCase)
        .filter(models.TestCase.id.in_(body.ids))
        .delete(synchronize_session=False)
    )
    db.commit()
    return BulkResult(affected=int(n))


@router.patch("/test-cases/bulk", response_model=BulkResult)
def bulk_update_test_cases(body: BulkUpdateRequest, db: Session = Depends(get_db)):
    patch = body.patch.model_dump(exclude_unset=True)
    if not patch:
        raise HTTPException(422, "Empty patch")
    n = (
        db.query(models.TestCase)
        .filter(models.TestCase.id.in_(body.ids))
        .update(patch, synchronize_session=False)
    )
    db.commit()
    return BulkResult(affected=int(n))


@router.get("/test-cases/{case_id}", response_model=TestCaseOut)
def get_test_case(case_id: int, db: Session = Depends(get_db)):
    obj = db.get(models.TestCase, case_id)
    if not obj:
        raise HTTPException(404, "TestCase not found")
    return obj


@router.patch("/test-cases/{case_id}", response_model=TestCaseOut)
def update_test_case(
    case_id: int, body: TestCaseUpdate, db: Session = Depends(get_db)
):
    obj = db.get(models.TestCase, case_id)
    if not obj:
        raise HTTPException(404, "TestCase not found")
    data = body.model_dump(exclude_unset=True)
    for k, v in data.items():
        setattr(obj, k, v)
    db.commit()
    db.refresh(obj)
    return obj


@router.delete("/test-cases/{case_id}", response_model=Msg)
def delete_test_case(case_id: int, db: Session = Depends(get_db)):
    obj = db.get(models.TestCase, case_id)
    if not obj:
        raise HTTPException(404, "TestCase not found")
    db.delete(obj)
    db.commit()
    return Msg(message="ok")


@router.get("/projects/{project_id}/stats", response_model=ProjectStats)
def project_stats(project_id: int, db: Session = Depends(get_db)):
    _project_or_404(db, project_id)
    doc_count = (
        db.query(func.count(models.Document.id))
        .filter(models.Document.project_id == project_id)
        .scalar()
    )
    case_q = db.query(models.TestCase).filter(
        models.TestCase.project_id == project_id
    )
    case_count = case_q.count()
    avg_score = (
        db.query(func.avg(models.TestCase.score))
        .filter(
            models.TestCase.project_id == project_id,
            models.TestCase.score.isnot(None),
        )
        .scalar()
    )
    type_rows = (
        db.query(models.TestCase.case_type, func.count(models.TestCase.id))
        .filter(models.TestCase.project_id == project_id)
        .group_by(models.TestCase.case_type)
        .all()
    )
    priority_rows = (
        db.query(models.TestCase.priority, func.count(models.TestCase.id))
        .filter(models.TestCase.project_id == project_id)
        .group_by(models.TestCase.priority)
        .all()
    )
    return ProjectStats(
        project_id=project_id,
        doc_count=int(doc_count or 0),
        case_count=case_count,
        avg_score=round(float(avg_score), 2) if avg_score is not None else None,
        type_dist={k or "unknown": v for k, v in type_rows},
        priority_dist={k or "unknown": v for k, v in priority_rows},
    )


@router.post("/test-cases/debug/graph")
def debug_judge_graph(body: JudgeGraphDebugRequest):
    """单步触发 judge StateGraph，返回最终用例、评分和每一步状态轨迹。

    用于直观查看『生成-评分-重生成』状态机的运行过程，不落库。
    """
    result = generate_with_judge_graph(
        body.point,
        body.case_type,
        max_retry=body.max_retry,
        threshold=body.threshold,
        return_trace=True,
    )
    return result


@router.get("/test-cases/debug/graph-mermaid")
def debug_judge_graph_mermaid():
    """返回 judge StateGraph 的 Mermaid 源码，可粘到 mermaid.live 预览。"""
    return {"mermaid": render_mermaid()}


@router.post("/test-cases/{case_id}/judge")
def rejudge_test_case(case_id: int, db: Session = Depends(get_db)):
    """对已保存的用例重新评分，并把最新分数写回数据库。"""
    obj = db.get(models.TestCase, case_id)
    if not obj:
        raise HTTPException(404, "TestCase not found")
    case_dict = {
        "title": obj.title,
        "preconditions": obj.preconditions,
        "steps": obj.steps,
        "expected": obj.expected,
        "priority": obj.priority,
    }
    score = judge_case(obj.title, obj.case_type, case_dict)
    obj.score = score["total"]
    db.commit()
    return {"case_id": obj.id, **score}
